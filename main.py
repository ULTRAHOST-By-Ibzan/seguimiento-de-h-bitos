import sys
from datetime import datetime, timedelta
import json
import sqlite3
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QCalendarWidget,
    QLabel,
    QHBoxLayout,
    QPushButton,
    QVBoxLayout,
    QLineEdit,
    QListWidget,
    QMessageBox,
    QInputDialog,
    QLCDNumber,
    QListWidgetItem,
    QProgressBar,
    QFileDialog,
    QDialog,
    QVBoxLayout,
    QLabel,
    QTableWidget,
    QTableWidgetItem
)
from PyQt5.QtCore import QDate, Qt, QTimer, QTime
from PyQt5 import QtGui
from PyQt5.QtGui import QTextCharFormat, QColor, QPixmap
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from os import path, remove
import openpyxl


class Calendar(QWidget):
    # keep the current time as class variable for reference
    currentDay = str(datetime.now().day).rjust(2, "0")
    currentMonth = str(datetime.now().month).rjust(2, "0")
    currentYear = str(datetime.now().year).rjust(2, "0")

    def __init__(self):
        super().__init__()
        folder = path.dirname(__file__)
        self.icon_folder = path.join(folder, "icons")

        self.setWindowTitle("hábitos  por IBZAN By ULTRAHOST")
        self.setWindowIcon(QtGui.QIcon(path.join(self.icon_folder, "window.png")))

        self.initUI()

        # Apply Fusion style
        app.setStyle("Fusion")

    def initUI(self):
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)

        # format for dates in calendar that have events
        self.fmt = QTextCharFormat()
        self.fmt.setBackground(QColor(255, 165, 0, 100))

        # format for the current day
        cur_day_fmt = QTextCharFormat()
        cur_day_fmt.setBackground(QColor(0, 255, 90, 70))

        # format to change back to if all events are deleted
        self.delfmt = QTextCharFormat()
        self.delfmt.setBackground(Qt.transparent)

        # SQLite database connection
        self.conn = sqlite3.connect('main.db')
        self.cursor = self.conn.cursor()
        self.create_table()

        # organize buttons and layouts for display
        self.addButton = QPushButton("+ Habito")
        self.editButton = QPushButton("Editar")
        self.delButton = QPushButton("Eliminar")
        self.reportButton = QPushButton("Reporte")
        self.viewButton = QPushButton("Ver Informe")

        self.note_group = QListWidget()
        self.note_group.setSortingEnabled(True)
        self.note_group.setStyleSheet("QListView::item {height: 40px;}")

        todayButton = QPushButton("Today")
        self.label = QLabel()

        labelp = QLabel()
        pixmap = QPixmap(path.join(self.icon_folder, "calendar.png"))
        labelp.setPixmap(pixmap)

        vbox = QVBoxLayout()
        vbox.addWidget(self.calendar)
        vbox.addWidget(self.note_group)
        vbox.addWidget(self.label)
        vbox.addWidget(labelp)

        hbox = QHBoxLayout()
        hbox.addLayout(vbox)
        buttons_layout = QVBoxLayout()
        buttons_layout.addWidget(self.addButton)
        buttons_layout.addWidget(self.editButton)
        buttons_layout.addWidget(self.delButton)
        buttons_layout.addWidget(self.reportButton)
        buttons_layout.addWidget(self.viewButton)
        hbox.addLayout(buttons_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("DIA: %p%")

        self.progress_bar_mes = QProgressBar()
        self.progress_bar_mes.setMinimum(0)
        self.progress_bar_mes.setMaximum(100)
        self.progress_bar_mes.setValue(0)
        self.progress_bar_mes.setFormat("MES: %p%")

        label_mes = QLabel("")
        label_mes.setAlignment(Qt.AlignCenter)

        vbox_bottom = QVBoxLayout()
        vbox_bottom.addLayout(hbox)
        vbox_bottom.addWidget(self.progress_bar)
        vbox_bottom.addWidget(self.progress_bar_mes)
        vbox_bottom.addWidget(label_mes)

        self.lcd = QLCDNumber()
        self.lcd.setDigitCount(8)
        self.lcd.setSegmentStyle(QLCDNumber.Filled)
        self.lcd.setStyleSheet("color: black; background-color: lightgrey; border: 2px solid black; border-radius: 5px;")

        vbox_bottom.addWidget(self.lcd)
        self.setLayout(vbox_bottom)

        # Connect signals and slots
        self.addButton.clicked.connect(self.addHabit)
        self.editButton.clicked.connect(self.editHabit)
        self.delButton.clicked.connect(self.deleteHabit)
        self.reportButton.clicked.connect(self.generate_report)
        self.viewButton.clicked.connect(self.view_monthly_report)
        self.calendar.selectionChanged.connect(self.showDateInfo)
        self.calendar.selectionChanged.connect(self.labelDate)
        self.calendar.selectionChanged.connect(self.highlightFirstItem)
        self.calendar.selectionChanged.connect(self.toggleAddEditDeleteButtons)
        self.note_group.itemChanged.connect(self.toggleHabitCompletion)

        # Show current month's habits
        self.showDateInfo()

        # Set up timer for the clock
        timer = QTimer(self)
        timer.timeout.connect(self.showTime)
        timer.start(1000)

    def create_table(self):
        self.cursor.execute('''CREATE TABLE IF NOT EXISTS habits
                             (date TEXT, habit TEXT, completed INTEGER)''')
        self.conn.commit()

    def addHabit(self):
        # Function to add habit
        habit_name, ok = QInputDialog.getText(self, "Agregar hábito", "Ingrese el nombre del hábito:")

        if ok and habit_name:
            start_date = datetime.now()
            end_date = start_date + timedelta(days=365)  # Un año completo

            current_date = start_date
            while current_date <= end_date:
                date = current_date.strftime("%d%m%Y")
                self.cursor.execute("INSERT INTO habits VALUES (?, ?, ?)", (date, habit_name, 0))
                self.conn.commit()
                current_date += timedelta(days=1)

            self.showDateInfo()

    def editHabit(self):
        # Function to edit habit
        item = self.note_group.currentItem()
        if item:
            new_habit_name, ok = QInputDialog.getText(self, "Editar hábito", "Ingrese el nombre del hábito:", text=item.text())
            if ok and new_habit_name:
                old_habit_name = item.text()
                self.cursor.execute("UPDATE habits SET habit=? WHERE habit=?", (new_habit_name, old_habit_name))
                self.conn.commit()
                self.showDateInfo()

    def deleteHabit(self):
        # Function to delete habit
        item = self.note_group.currentItem()
        if item:
            habit_name = item.text()
            reply = QMessageBox.question(self, "Eliminar Habit", f"¿Eliminar el habito '{habit_name}'?", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.cursor.execute("DELETE FROM habits WHERE habit=?", (habit_name,))
                self.conn.commit()
                self.showDateInfo()

    def toggleHabitCompletion(self, item):
        # Function to toggle habit completion
        habit_name = item.text()
        completed = 1 if item.checkState() == Qt.Checked else 0
        date = self.calendar.selectedDate().toString("ddMMyyyy")
        self.cursor.execute("UPDATE habits SET completed=? WHERE habit=? AND date=?", (completed, habit_name, date))
        self.conn.commit()
        self.updateProgressBar()

    def showDateInfo(self):
        # Function to display habits for selected date
        self.note_group.clear()
        date = self.calendar.selectedDate().toString("ddMMyyyy")
        self.cursor.execute("SELECT habit, completed FROM habits WHERE date=?", (date,))
        habits = self.cursor.fetchall()
        for habit in habits:
            habit_name, completed = habit
            item = QListWidgetItem(habit_name)
            if completed:
                item.setCheckState(Qt.Checked)
                item.setBackground(Qt.green)  # Set green background for completed habits
            else:
                item.setCheckState(Qt.Unchecked)
            self.note_group.addItem(item)
        self.updateProgressBar()

    def selectToday(self):
        self.calendar.setSelectedDate(QDate.currentDate())

    def toggleAddEditDeleteButtons(self):
        # Enable/disable buttons based on selection
        enabled = self.calendar.selectedDate() >= QDate.currentDate()
        for button in [self.addButton, self.editButton, self.delButton, self.reportButton, self.viewButton]:
            button.setEnabled(enabled)

    def labelDate(self):
        # Set label to show the long name form of the selected date
        select = self.calendar.selectedDate()
        weekday, month = select.dayOfWeek(), select.month()
        day, year = str(select.day()), str(select.year())
        week_day, word_month = QDate.longDayName(weekday), QDate.longMonthName(month)
        self.label.setText(week_day + ", " + word_month + " " + day + ", " + year)

    def highlightFirstItem(self):
        # Highlight the first item immediately after switching selection
        if self.note_group.count() > 0:
            self.note_group.setCurrentRow(0)

    def showTime(self):
        # Keep the current time updated
        time = QTime.currentTime()
        text = time.toString("hh:mm:ss")
        self.lcd.display(text)

    def updateProgressBar(self):
        # Update progress bar based on completion status of habits
        total_habits = self.note_group.count()
        completed_habits = sum(1 for i in range(total_habits) if self.note_group.item(i).checkState() == Qt.Checked)
        if total_habits > 0:
            percentage = (completed_habits / total_habits) * 100
            self.progress_bar.setValue(int(percentage))
        else:
            self.progress_bar.setValue(0)

        # Update monthly progress bar
        current_date = datetime.now()
        start_date = current_date.replace(day=1)
        end_date = current_date.replace(day=1, month=current_date.month+1) - timedelta(days=1)
        days_in_month = (end_date - start_date).days + 1
        completed_habits_month = 0

        current_date = start_date
        while current_date <= end_date:
            date = current_date.strftime("%d%m%Y")
            self.cursor.execute("SELECT completed FROM habits WHERE date=?", (date,))
            habits = self.cursor.fetchall()
            for habit in habits:
                completed_habits_month += habit[0]
            current_date += timedelta(days=1)

        if total_habits > 0 and days_in_month > 0:
            percentage_month = (completed_habits_month / (total_habits * days_in_month)) * 100
            self.progress_bar_mes.setValue(int(percentage_month))
        else:
            self.progress_bar_mes.setValue(0)

    def generate_report(self):
        options = ("Mes en Curso", "Último Mes", "Último Trimestre")
        choice, ok = QInputDialog.getItem(self, "Seleccionar Reporte", "Selecciona el tipo de reporte:", options, 0, False)
        if ok and choice:
            if choice == "Mes en Curso":
                self.generate_monthly_report()
            elif choice == "Último Mes":
                self.generate_last_month_report()
            elif choice == "Último Trimestre":
                self.generate_last_quarter_report()

    def generate_monthly_report(self):
        current_date = datetime.now()
        start_date = current_date.replace(day=1)
        end_date = current_date.replace(day=1, month=current_date.month+1) - timedelta(days=1)
        filename = "monthly_report.xlsx"
        self.generate_report_by_date_range(start_date, end_date, filename)

    def generate_last_month_report(self):
        current_date = datetime.now()
        start_date = current_date.replace(day=1, month=current_date.month-1)
        end_date = current_date.replace(day=1) - timedelta(days=1)
        filename = "ultimo-mes.xlsx"
        self.generate_report_by_date_range(start_date, end_date, filename)

    def generate_last_quarter_report(self):
        current_date = datetime.now()
        start_date = current_date - timedelta(days=90)
        end_date = current_date
        filename = "ultimo-trimestre.xlsx"
        self.generate_report_by_date_range(start_date, end_date, filename)

    def generate_report_by_date_range(self, start_date, end_date, filename):
        data = {}
        current_date = start_date
        while current_date <= end_date:
            formatted_date = current_date.strftime("%d/%m/%Y")  # Formato DD/MM/YYYY
            habits = self.fetch_habits(current_date.strftime("%d%m%Y"))
            for habit in habits:
                habit_name = habit[1]
                completed = "Sí" if habit[2] == 1 else "No"
                if habit_name not in data:
                    data[habit_name] = {"total": 0, "completed": 0}  # Inicializar conteo de hábitos
                data[habit_name]["total"] += 1
                if completed == "Sí":
                    data[habit_name]["completed"] += 1
                if formatted_date not in data[habit_name]:
                    data[habit_name][formatted_date] = completed
            current_date += timedelta(days=1)

        try:
            wb = Workbook()
            ws = wb.active

            # Write header row with dates
            header_row = ["Fecha"]
            dates = sorted(list(set(date for habit_data in data.values() for date in habit_data.keys())))
            header_row.extend(dates)
            header_row.append("Porcentaje Completado")
            ws.append(header_row)

            # Write habit rows
            for habit_name, habit_data in data.items():
                row = [habit_name]
                for date in dates:
                    row.append(habit_data.get(date, "-"))
                total_habits = habit_data["total"]
                completed_habits = habit_data["completed"]
                percentage_completed = f"{(completed_habits / total_habits) * 100:.2f}%" if total_habits > 0 else "0%"
                row.append(percentage_completed)
                ws.append(row)

            # Apply conditional formatting
            for row in ws.iter_rows(min_row=2, max_col=len(header_row), max_row=len(data) + 1):
                for cell in row:
                    if cell.value == "Sí":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                                fill_type="solid")  # Verde
                    elif cell.value == "No":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo

            wb.save(filename)
            self.show_message_box(f"Informe guardado como {filename}")

        except Exception as e:
            self.show_message_box(f"Error al generar el informe: {e}")

    def fetch_habits(self, date):
        self.cursor.execute("SELECT * FROM habits WHERE date=?", (date,))
        return self.cursor.fetchall()

    def view_monthly_report(self):
        current_date = datetime.now()
        start_date = current_date.replace(day=1)
        end_date = current_date.replace(day=1, month=current_date.month+1) - timedelta(days=1)
        filename = "reporte-mensual.xlsx"
        self.generate_report_by_date_range(start_date, end_date, filename)
        self.view_excel_report(filename)

    def view_excel_report(self, filename):
        dialog = QDialog(self)
        dialog.setWindowTitle("Informe Mensual")
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.resize(1200, 700)
        layout = QVBoxLayout()

        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active

            table = QTableWidget()
            table.setRowCount(sheet.max_row)
            table.setColumnCount(sheet.max_column)

            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    table.setItem(row - 1, col - 1, QTableWidgetItem(str(cell.value)))

            layout.addWidget(table)

            dialog.setLayout(layout)
            dialog.exec_()
            wb.close()
        except Exception as e:
            self.show_message_box(f"Error al abrir el informe: {e}")

        try:
            remove(filename)
        except Exception as e:
            self.show_message_box(f"Error al eliminar el archivo temporal: {e}")

    def show_message_box(self, message):
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Mensaje")
        msg_box.setText(message)
        msg_box.exec_()

    def closeEvent(self, e):
        # Close the database connection
        self.conn.close()
        e.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Calendar()
    window.show()
    sys.exit(app.exec_())
